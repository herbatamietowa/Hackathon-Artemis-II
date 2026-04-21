"""FastAPI route handlers."""
from __future__ import annotations
import logging
import re
from datetime import date

import pandas as pd
from pydantic import BaseModel
from fastapi import APIRouter, File, HTTPException, UploadFile

from .schemas import (
    AgentTurn,
    UploadDataResponse,
    AnalyzeRequest,
    AnalyzeResponse,
    ApproveProjectRequest,
    ConfirmProjectRequest,
    DebateProjectPathRequest,
    DebateProjectPathResponse,
    DisasterRequest,
    DisasterResult,
    FactoryListResponse,
    GCIRequest,
    GCIResponse,
    GCIRouteModel,
    MaterialListResponse,
    ProjectArchitectRequest,
    ProjectArchitectResponse,
    ProjectSimulationRequest,
    ProjectSimulationResponse,
    RawMaterialItem,
    RawMaterialListResponse,
    RawMaterialOrderRequest,
    RawMaterialStatusModel,
    ReallocationSuggestion,
    ScenarioListResponse,
    ScenarioPathModel,
    SimulationPathModel,
    SourcingRequest,
    SourcingResponse,
    WCLoadModel,
)
from ..config import DATA_PATH, SCENARIOS
from ..data.loader import load_workbook
from ..engine.capacity import compute_capacity_plan

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api")


class TimelinePoint(BaseModel):
    period: str
    capacity_utilization: float
    available_hours: float
    demanded_hours: float
    bottleneck_detected: bool


class TimelineResponse(BaseModel):
    factory: str
    scenario: str
    points: list[TimelinePoint]


@router.get("/health")
def health() -> dict:
    return {"status": "ok"}


@router.get("/scenarios", response_model=ScenarioListResponse)
def list_scenarios() -> ScenarioListResponse:
    return ScenarioListResponse(scenarios=SCENARIOS)


@router.get("/factories", response_model=FactoryListResponse)
def list_factories() -> FactoryListResponse:
    try:
        wb = load_workbook(DATA_PATH)
        df = wb.get("2_1 Work Center Capacity Weekly", pd.DataFrame())
        codes = df["Work center code"].dropna().astype(str)
        factories = sorted({m.group(1) for wc in codes for m in [re.search(r"(NW\d+)", wc)] if m})
        if factories:
            return FactoryListResponse(factories=factories)
    except Exception as exc:
        logger.warning("Could not derive factory list from data: %s", exc)
    return FactoryListResponse(factories=["NW01", "NW03"])


@router.get("/timeline", response_model=TimelineResponse)
def timeline(
    factory: str = "NW01",
    scenario: str = "probability_weighted",
    months: int = 36,
) -> TimelineResponse:
    """Return capacity utilization for up to 36 months."""
    if scenario not in SCENARIOS:
        raise HTTPException(status_code=400, detail=f"Unknown scenario: {scenario}")
    months = min(max(months, 1), 36)

    today = date.today()
    periods: list[str] = []
    m, y = today.month, today.year
    for _ in range(months):
        m += 1
        if m > 12:
            m, y = 1, y + 1
        periods.append(f"{m} {y}")

    points: list[TimelinePoint] = []
    for period in periods:
        try:
            r = compute_capacity_plan(factory, scenario, period, DATA_PATH)
            points.append(TimelinePoint(
                period=period,
                capacity_utilization=r.capacity_utilization,
                available_hours=r.available_hours,
                demanded_hours=r.demanded_hours,
                bottleneck_detected=r.bottleneck_detected,
            ))
        except Exception as exc:
            logger.warning("Timeline skipping period %s: %s", period, exc)

    return TimelineResponse(factory=factory, scenario=scenario, points=points)


@router.post("/analyze", response_model=AnalyzeResponse)
def analyze(req: AnalyzeRequest) -> AnalyzeResponse:
    if req.scenario not in SCENARIOS:
        raise HTTPException(status_code=400, detail=f"Unknown scenario: {req.scenario}. Valid: {SCENARIOS}")

    period = req.period
    if period is None:
        today = date.today()
        month = today.month % 12 + 1
        year = today.year + (1 if today.month == 12 else 0)
        period = f"{month} {year}"

    try:
        from ..agents.agent1_capacity import run_agent1
        from ..agents.agent2_sustainability import run_agent2

        engine_result = compute_capacity_plan(req.factory, req.scenario, period, DATA_PATH)

        debate_history: list[AgentTurn] = []
        if req.user_argument:
            debate_history.append(AgentTurn(agent_name="User", message=req.user_argument))

        agent1_result = run_agent1(engine_result, debate_context=[t.model_dump() for t in debate_history] if debate_history else None)
        debate_history.append(AgentTurn(agent_name="Cost Specialist", message=agent1_result.reasoning))

        agent2_verdict = run_agent2(agent1_result, debate_context=[t.model_dump() for t in debate_history])
        debate_history.append(AgentTurn(agent_name="Sustainability Director", message=agent2_verdict.strategy, verdict=agent2_verdict.verdict))

        if agent2_verdict.verdict == "REOPEN DEBATE":
            agent1_result = run_agent1(engine_result, debate_context=[t.model_dump() for t in debate_history])
            debate_history.append(AgentTurn(agent_name="Cost Specialist", message=agent1_result.reasoning))
            agent2_verdict = run_agent2(agent1_result, debate_context=[t.model_dump() for t in debate_history])
            debate_history.append(AgentTurn(agent_name="Sustainability Director", message=agent2_verdict.strategy, verdict=agent2_verdict.verdict))

        status = "USER_OVERRIDE" if req.user_argument else ("CONSENSUS" if agent2_verdict.verdict == "APPROVED" else "CONTESTED")

        # Reallocation suggestion — only when bottleneck and factory is not already NW03
        reallocation: ReallocationSuggestion | None = None
        if engine_result.bottleneck_detected and req.factory != "NW03":
            from ..engine.reallocation import check_nw03_headroom
            overflow = max(0.0, engine_result.demanded_hours - engine_result.available_hours)
            reallocation = check_nw03_headroom(period, overflow, DATA_PATH, source_factory=req.factory)

        return AnalyzeResponse(
            agent1_result=agent1_result,
            agent2_verdict=agent2_verdict,
            per_work_center=[
                WCLoadModel(
                    wc=wc.wc,
                    utilization=wc.utilization,
                    available=wc.available,
                    demanded=wc.demanded,
                )
                for wc in engine_result.per_work_center
            ],
            debate_history=debate_history,
            status=status,
            reallocation=reallocation,
        )
    except Exception as exc:
        logger.exception("analyze endpoint error: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/sourcing", response_model=SourcingResponse)
def sourcing(req: SourcingRequest) -> SourcingResponse:
    if req.scenario not in SCENARIOS:
        raise HTTPException(status_code=400, detail=f"Unknown scenario: {req.scenario}")
    try:
        from ..engine.sourcing import compute_sourcing_plan
        return compute_sourcing_plan(req.factory, req.scenario, req.period, DATA_PATH)
    except Exception as exc:
        logger.exception("sourcing endpoint error: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/materials", response_model=MaterialListResponse)
def list_materials() -> MaterialListResponse:
    """Return all active materials that have multi-plant tooling (GCI candidates)."""
    try:
        wb = load_workbook(DATA_PATH)
        df26 = wb.get("2_6 Tool_material nr master", pd.DataFrame())
        df23 = wb.get("2_3 SAP MasterData", pd.DataFrame())
        active = df26[df26["Material Status"] == "Active"]
        multi = active.groupby("Sap code")["Plant"].nunique()
        multi_codes = multi[multi > 1].index.tolist()
        name_map = df23.set_index("Sap code")["Description"].to_dict()
        materials = [
            {"code": code, "name": name_map.get(code, code)}
            for code in sorted(multi_codes)
        ]
        return MaterialListResponse(materials=materials)
    except Exception as exc:
        logger.warning("Could not load materials: %s", exc)
        return MaterialListResponse(materials=[])


@router.post("/gci", response_model=GCIResponse)
def gci(req: GCIRequest) -> GCIResponse:
    try:
        from ..engine.gci import compute_gci
        from ..agents.agent2_sustainability import run_agent2_gci

        result = compute_gci(
            material_code=req.material_code,
            rdd_str=req.rdd,
            alpha=req.alpha,
            data_path=DATA_PATH,
            forced_mode=req.forced_mode,
        )

        ai_insight = run_agent2_gci(result.ai_context)

        routes = [
            GCIRouteModel(
                plant=r.plant,
                plant_name=r.plant_name,
                region=r.region,
                mode=r.mode,
                gci=round(r.gci, 4),
                cost_score=round(r.cost_score, 4),
                carbon_score=round(r.carbon_score, 4),
                raw_cost_eur=round(r.raw_cost_eur, 2),
                raw_carbon=round(r.raw_carbon, 4),
                grid_intensity=r.grid_intensity,
                scrap_factor=round(r.scrap_factor, 4),
                dominant_size=r.dominant_size,
                arrival_date=r.arrival_date.isoformat(),
                meets_rdd=r.meets_rdd,
                days_margin=r.days_margin,
                transport_lt_days=r.transport_lt_days,
                carbon_penalty=r.carbon_penalty,
            )
            for r in result.routes
        ]

        return GCIResponse(
            material_code=result.material_code,
            material_name=result.material_name,
            rdd=req.rdd,
            slider_alpha=req.alpha,
            forced_mode=req.forced_mode,
            routes=routes,
            recommended_plant=result.recommended.plant if result.recommended else None,
            green_baseline=round(result.green_baseline, 4),
            green_potential_saving_pct=round(result.green_potential_saving_pct, 1),
            ai_insight=ai_insight,
        )
    except Exception as exc:
        logger.exception("GCI endpoint error: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/disaster", response_model=DisasterResult)
def disaster(req: DisasterRequest) -> DisasterResult:
    if req.scenario not in SCENARIOS:
        raise HTTPException(status_code=400, detail=f"Unknown scenario: {req.scenario}")
    if not (1 <= req.duration_months <= 12):
        raise HTTPException(status_code=400, detail="duration_months must be between 1 and 12")
    try:
        from ..engine.disaster import compute_disaster_impact
        period = req.period
        if period is None:
            today = date.today()
            month = today.month % 12 + 1
            year = today.year + (1 if today.month == 12 else 0)
            period = f"{month} {year}"
        return compute_disaster_impact(
            offline_factory=req.offline_factory,
            scenario=req.scenario,
            period=period,
            duration_months=req.duration_months,
            data_path=DATA_PATH,
        )
    except Exception as exc:
        logger.exception("disaster endpoint error: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/project-architect", response_model=ProjectArchitectResponse)
def project_architect(req: ProjectArchitectRequest) -> ProjectArchitectResponse:
    try:
        from ..engine.project_architect import compute_project_architect
        result = compute_project_architect(
            material_code=req.material_code,
            quantity=req.quantity,
            deadline=req.deadline,
            data_path=DATA_PATH,
        )
        return ProjectArchitectResponse(
            material_code=result.material_code,
            material_name=result.material_name,
            quantity=result.quantity,
            deadline=result.deadline,
            paths=[ScenarioPathModel(**vars(p)) for p in result.paths],
        )
    except Exception as exc:
        logger.exception("project-architect error: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/confirm-project")
def confirm_project(req: ConfirmProjectRequest) -> dict:
    import csv
    from datetime import datetime
    log_path = DATA_PATH.parent / "confirmed_projects.csv"
    row = {
        "timestamp": datetime.utcnow().isoformat(),
        "material_code": req.material_code,
        "material_name": req.material_name,
        "quantity": req.quantity,
        "deadline": req.deadline or "",
        "chosen_path": req.chosen_path,
        "chosen_plant": req.chosen_plant,
        "cost_eur": req.cost_eur,
        "delivery_date": req.delivery_date,
    }
    try:
        write_header = not log_path.exists()
        with open(log_path, "a", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(row.keys()))
            if write_header:
                writer.writeheader()
            writer.writerow(row)
    except Exception as exc:
        logger.warning("Could not write project log: %s", exc)
    return {"status": "saved", "project": row}


@router.get("/plates", response_model=MaterialListResponse)
def list_plates() -> MaterialListResponse:
    """Return plate materials from 1_1 Export Plates."""
    try:
        from ..engine.project_simulation import get_plate_list
        return MaterialListResponse(materials=get_plate_list(DATA_PATH))
    except Exception as exc:
        logger.warning("Could not load plates: %s", exc)
        return MaterialListResponse(materials=[])


@router.get("/gaskets", response_model=MaterialListResponse)
def list_gaskets() -> MaterialListResponse:
    """Return gasket materials from 1_2 Gaskets."""
    try:
        from ..engine.project_simulation import get_gasket_list
        return MaterialListResponse(materials=get_gasket_list(DATA_PATH))
    except Exception as exc:
        logger.warning("Could not load gaskets: %s", exc)
        return MaterialListResponse(materials=[])


@router.post("/simulate-project", response_model=ProjectSimulationResponse)
def simulate_project(req: ProjectSimulationRequest) -> ProjectSimulationResponse:
    try:
        from ..engine.project_simulation import compute_project_simulation
        result = compute_project_simulation(
            plate_code=req.plate_code,
            quantity=req.quantity,
            data_path=DATA_PATH,
        )
        return ProjectSimulationResponse(
            plate_code=result.plate_code,
            plate_name=result.plate_name,
            gasket_code=result.gasket_code,
            gasket_name=result.gasket_name,
            quantity=result.quantity,
            feasible_plants=result.feasible_plants,
            raw_materials=[RawMaterialStatusModel(**vars(r)) for r in result.raw_materials],
            paths=[SimulationPathModel(**vars(p)) for p in result.paths],
            warning=result.warning,
        )
    except Exception as exc:
        logger.exception("simulate-project error: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/raw-materials", response_model=RawMaterialListResponse)
def list_raw_materials() -> RawMaterialListResponse:
    """Return component (raw) materials from BOM with current stock totals."""
    try:
        from collections import defaultdict
        wb = load_workbook(DATA_PATH)
        df32 = wb.get("3_2 Component_SF_RM", pd.DataFrame())
        df31 = wb.get("3_1 Inventory ATP", pd.DataFrame())
        df23 = wb.get("2_3 SAP MasterData", pd.DataFrame())

        rm_df = (
            df32[["Component Material code", "Component Description", "Component BUoM"]]
            .rename(columns={
                "Component Material code": "code",
                "Component Description": "name",
                "Component BUoM": "unit",
            })
            .drop_duplicates("code")
            .dropna(subset=["code"])
        )
        rm_df["code"] = rm_df["code"].astype(str).str.strip()
        rm_df["name"] = rm_df["name"].fillna(rm_df["code"]).astype(str)
        rm_df["unit"] = rm_df["unit"].fillna("PC").astype(str)
        rm_df = rm_df[rm_df["code"] != ""]

        stock_col = next((c for c in df31.columns if "Material Unique" in c and "code" in c.lower()), None)
        stock_map: dict = {}
        if stock_col and "Stock Qty" in df31.columns:
            stock_map = df31.groupby(stock_col)["Stock Qty"].sum().to_dict()

        # Derive RM unit cost: fg_standard_cost / effective_qty_per_pc, averaged across all FG mappings
        cost_map: dict = {}
        if "Sap code" in df23.columns and "Standard Cost in EUR" in df23.columns:
            cost_map = df23.dropna(subset=["Standard Cost in EUR"]).set_index("Sap code")["Standard Cost in EUR"].to_dict()

        rm_costs: dict = defaultdict(list)
        for _, bom_row in df32.iterrows():
            rm_code = str(bom_row.get("Component Material code", "")).strip()
            fg_code = str(bom_row.get("Header Material code", "")).strip()
            eff_qty = float(bom_row.get("Effective Component Quantity", 0) or 0)
            fg_cost = cost_map.get(fg_code, 0.0)
            if rm_code and eff_qty > 0 and fg_cost > 0:
                rm_costs[rm_code].append(fg_cost / eff_qty)

        materials = [
            RawMaterialItem(
                code=row["code"],
                name=row["name"],
                unit=row["unit"],
                stock_qty=float(stock_map.get(row["code"], 0.0)),
                unit_cost_eur=round(sum(rm_costs[row["code"]]) / len(rm_costs[row["code"]]), 4) if rm_costs[row["code"]] else None,
            )
            for _, row in rm_df.sort_values("code").iterrows()
        ]
        return RawMaterialListResponse(materials=materials)
    except Exception as exc:
        logger.warning("Could not load raw materials: %s", exc)
        return RawMaterialListResponse(materials=[])


@router.post("/order-raw-material")
def order_raw_material(req: RawMaterialOrderRequest) -> dict:
    import csv
    from datetime import datetime
    log_path = DATA_PATH.parent / "raw_material_orders.csv"
    order_id = f"RM-{int(datetime.utcnow().timestamp())}"
    row = {
        "order_id": order_id,
        "timestamp": datetime.utcnow().isoformat(),
        "material_code": req.material_code,
        "material_name": req.material_name,
        "unit": req.unit,
        "quantity": req.quantity,
        "factory": req.factory,
        "deadline": req.deadline or "",
    }
    try:
        write_header = not log_path.exists()
        with open(log_path, "a", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(row.keys()))
            if write_header:
                writer.writeheader()
            writer.writerow(row)
    except Exception as exc:
        logger.warning("Could not write raw material order: %s", exc)
    return {"status": "ordered", "order_id": order_id}


@router.post("/debate-project-path", response_model=DebateProjectPathResponse)
def debate_project_path(req: DebateProjectPathRequest) -> DebateProjectPathResponse:
    try:
        from ..engine.project_simulation import compute_project_simulation
        from ..agents.path_debate import run_path_debate

        result = compute_project_simulation(
            plate_code=req.plate_code,
            quantity=req.quantity,
            data_path=DATA_PATH,
        )

        if not result.paths:
            raise HTTPException(status_code=400, detail="No feasible paths found for this material.")

        paths_dicts = [vars(p) for p in result.paths]
        debate = run_path_debate(
            paths=paths_dicts,
            plate_code=result.plate_code,
            plate_name=result.plate_name,
            user_argument=req.user_argument,
        )

        return DebateProjectPathResponse(
            agreed_path=SimulationPathModel(**debate["agreed_path"]),
            debate_history=[AgentTurn(**t) for t in debate["debate_history"]],
            status=debate["status"],
            parameters_considered=debate["parameters_considered"],
            tradeoffs=debate.get("tradeoffs", []),
            plate_code=result.plate_code,
            plate_name=result.plate_name,
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("debate-project-path error: %s", exc)
        raise HTTPException(status_code=500, detail=str(exc))


@router.post("/approve-project")
def approve_project(req: ApproveProjectRequest) -> dict:
    import json
    from datetime import datetime
    log_path = DATA_PATH.parent / "approved_projects.jsonl"
    record = {
        "project_id": f"PROJ-{int(datetime.utcnow().timestamp())}",
        "project_name": f"{req.plate_code} × {req.quantity}",
        "plate_code": req.plate_code,
        "plate_name": req.plate_name,
        "gasket_code": req.gasket_code,
        "quantity": req.quantity,
        "status": "Approved",
        "plant": req.plant,
        "mode": req.mode,
        "note": f"{req.mode} logistics — {req.path_name} scenario",
        "total_cost_eur": req.total_cost_eur,
        "delivery_days": req.delivery_days,
        "carbon_score": req.carbon_score,
        "approved_at": datetime.utcnow().isoformat(),
    }
    try:
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(json.dumps(record) + "\n")
    except Exception as exc:
        logger.warning("Could not write approved project: %s", exc)
    return {"status": "approved", "record": record}


@router.post("/upload-data", response_model=UploadDataResponse)
async def upload_data(file: UploadFile = File(...)) -> UploadDataResponse:
    """Merge an uploaded .xlsx into the main dataset and invalidate the cache."""
    import io
    import openpyxl
    from ..data.loader import SHEET_NAMES, invalidate_cache

    contents = await file.read()
    try:
        uploaded = pd.read_excel(io.BytesIO(contents), sheet_name=None, engine="openpyxl")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Cannot read uploaded file: {exc}")

    uploaded_norm = {k.strip(): v for k, v in uploaded.items()}
    known = {s.strip() for s in SHEET_NAMES}

    # Match uploaded sheets to known sheets by exact name or 3-char prefix
    matches: dict[str, pd.DataFrame] = {}
    for up_name, up_df in uploaded_norm.items():
        target = next((k for k in known if up_name == k or up_name[:3] == k[:3]), None)
        if target is not None:
            matches[target] = up_df

    if not matches:
        raise HTTPException(status_code=400, detail="No sheets matched known dataset sheets. Check your file.")

    # Load the workbook with openpyxl and append rows directly — avoids full read/rewrite
    try:
        wb = openpyxl.load_workbook(DATA_PATH)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Cannot read dataset: {exc}")

    rows_added: dict[str, int] = {}
    for target, up_df in matches.items():
        ws = next((wb[s] for s in wb.sheetnames if s.strip() == target), None)
        if ws is None:
            continue
        headers = [cell.value for cell in ws[1]]
        for _, row in up_df.iterrows():
            new_row = []
            for h in headers:
                val = row.get(h) if h in up_df.columns else None
                new_row.append(None if (val is None or (isinstance(val, float) and pd.isna(val))) else val)
            ws.append(new_row)
        rows_added[target] = len(up_df)

    if not rows_added:
        raise HTTPException(status_code=400, detail="Sheets matched but no rows could be appended.")

    try:
        wb.save(DATA_PATH)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to save dataset: {exc}")

    invalidate_cache()
    return UploadDataResponse(sheets_merged=list(rows_added.keys()), rows_added=rows_added)


@router.get("/sample-factory-data")
def sample_factory_data():
    """Return a sample .xlsx for a new factory NW16 that can be uploaded via /upload-data."""
    import io
    from fastapi.responses import StreamingResponse

    # Weeks covering 2026 (same format the calendar engine uses)
    weeks = [f"Week {w} 2026" for w in range(1, 53)]
    wcs = ["PRESSING", "WELDING", "ASSEMBLY"]

    # 2_1: Work Center Capacity Weekly — 120h per week per WC
    rows_21 = []
    for wc in wcs:
        row: dict = {"Work center code": f"P01_NW16_{wc}", "Measure": "Available Capacity, hours"}
        for week in weeks:
            row[week] = 120.0
        rows_21.append(row)

    # 2_5: WC Schedule limits (OEE)
    rows_25 = [
        {"Plant": "NW16", "Plant name": "Nordic Hub", "WC-Description": wc,
         "OEE (in %)": 0.85, "AP Limit": "Available Capacity, hours"}
        for wc in wcs
    ]

    # 2_6: Tool master — a few mock materials active at NW16
    mat_codes = [f"NW16-MAT-{i:04d}" for i in range(1, 6)]
    rows_26 = []
    for mat in mat_codes:
        rows_26.append({
            "Sap code": mat, "Plant": "NW16", "Material Status": "Active",
            "Work center": "PRESSING", "Cycle time": 2.5,
            "Connector Plant_Material nr": f"NW16_{mat}",
            "Cycle times Standard Value (Machine)": 2.5, "Rev no": 1,
        })

    # 2_3: SAP MasterData for the mock materials
    rows_23 = [
        {"Sap code": mat, "Description": f"NW16 Test Part {mat}",
         "Standard Cost in EUR": 180.0, "G35 - Plant": "NW16"}
        for mat in mat_codes
    ]

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        pd.DataFrame(rows_21).to_excel(writer, sheet_name="2_1 Work Center Capacity Weekly", index=False)
        pd.DataFrame(rows_25).to_excel(writer, sheet_name="2_5 WC Schedule_limits", index=False)
        pd.DataFrame(rows_26).to_excel(writer, sheet_name="2_6 Tool_material nr master", index=False)
        pd.DataFrame(rows_23).to_excel(writer, sheet_name="2_3 SAP MasterData", index=False)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=mock_factory_NW16.xlsx"},
    )
